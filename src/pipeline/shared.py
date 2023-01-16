from __future__ import annotations

import os
import pprint
from abc import ABC
from abc import abstractmethod
from itertools import combinations
from itertools import islice
from typing import Any
from typing import Callable
from typing import cast
from typing import Generator
from typing import Literal
from typing import Sequence
from typing import TYPE_CHECKING
from typing import TypeVar

import numpy as np
import pandas as pd
from joblib import delayed
from joblib import dump
from joblib import load
from joblib import Parallel
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame
from pandas import ExcelWriter
from pandas import Index
from scipy.stats import ttest_ind
from scipy.stats import ttest_rel
from sklearn.ensemble import RandomForestClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.neural_network import MLPClassifier
from sklearn.svm import SVC
from sklearn.tree import DecisionTreeClassifier
from src.models.decision_tree import DecisionTree
from statsmodels.stats.multitest import fdrcorrection
from xgboost import XGBClassifier

from .. import data_models
from .._types import BaseEstimator
from .._types import KArgs
from .._types import KDEType
from .._types import Models
from .._types import Pipelines
from .._types import Proteins
from .._types import T
from ..mccv import mc_cv
from ..models.logisitic_regression import LogisticRegression as LogReg
from ..models.multi_layer_perceptron import MultiLayerPerceptron
from ..models.random_forest import RandomForest
from ..models.support_vector_machine import SupportVectorMachine
from ..models.xgboost import XGBoost
from ..parameters import Parameter

if TYPE_CHECKING:
    from .biological import Biological
    from .non_biological import NonBiological
    from .._types import ModelEstimator


pp = pprint.PrettyPrinter(indent=2, width=200, compact=True, sort_dicts=False)


class FunctionNotCompatibleError(Exception):
    """Error for incompatible pipelines."""

    def __init__(self, name: str, message: str | None = None) -> None:
        self.name = name
        self.message = f"'{self.name}' is not compatible with the chosen pipeline." if message is None else message
        super().__init__(self.message)


RT = TypeVar("RT")


def pipeline(name: T) -> Callable[[Callable[..., RT]], Callable[..., RT]]:
    """A decorator function for constricting pipelines.

    Parameters
    ----------
    name
        The name of the pipeline

    Raises
    ------
    FunctionNotCompatibleError
        If `name` is not the same as the classes pipeline name.

    Returns
    -------
        The called function.
    """

    def decorator(func: Callable[..., RT]) -> Callable[..., RT]:
        def wrapper(self: Biological | NonBiological, *args: KArgs.args, **kwargs: KArgs.kwargs) -> RT:
            if name != self.pipeline:
                raise FunctionNotCompatibleError(func.__name__)
            ret = func(self, *args, **kwargs)
            return ret

        return wrapper

    return decorator


class SharedPipeline(ABC):
    """A collective class of shared pipeline methods."""

    def __init__(self) -> None:
        self.kde_methods: Sequence[KDEType] = [
            "balanced",
            "imbalanced_early",
            "imbalanced_healthy",
            "none",
        ]
        self.constant_options: data_models.ConstantOptions
        self.combo_options: data_models.ComboOptions
        self.hyperparameters: Sequence[Parameter]
        self.pipeline: Pipelines
        self.model_name: Models
        self.collection: str
        self.kde_method: KDEType
        self._collections: Sequence[str] = [""]

    @property
    def file_names(self) -> list[str]:
        """A collection of filenames.

        Returns
        -------
            A list of filenames for each classifier.
        """
        return [
            "multi_layer_perceptron",
            "logistic_regression_l2",
            "support_vector_machine",
            "random_forest",
            "decision_tree",
            "xgboost",
        ]

    @property
    def param_names(self) -> list[list[str]]:
        """A collection of classifier parameter names.

        Returns
        -------
            A 2D list of parameter names for each classifier.
        """
        return [
            ["alpha", "hidden_layer_sizes", "early_stopping"],
            ["penalty", "C", "solver"],
            ["C", "gamma", "kernel"],
            ["n_estimators", "max_features", "max_depth", "criterion"],
            ["max_features", "max_depth", "criterion", "splitter"],
            ["n_estimators", "max_depth", "learning_rate", "reg_lambda"],
        ]

    @property
    def collections(self) -> Sequence[str]:
        """A collection of collections.

        Returns
        -------
            A custom sequence of collections.
        """
        return self._collections

    @collections.setter
    def collections(self, value: Sequence[str]) -> None:
        self._collections = value

    def get_best_hyperparameters(
        self,
        method: KDEType,
        /,
        control: bool = False,
        biological: bool = False,
        version: int | str = 2,
    ) -> dict[str, dict]:
        """Return the best hyperparameter values from all models.

        Parameters
        ----------
        method
            The Kernel Density method to use
        control
            Whether or not control data is being used, by default False
            Will skip "DP" collection if True.
        biological
            Toggle to adjust when using biological data, by default False
        version
            The version identifier for a set of hyperparameters,
            by default 2

        Returns
        -------
            A dictionary, ordered by collection and model with the
            best parameters.
        """
        file_names = self.file_names
        param_names = self.param_names
        folder = self.combo_options.pipeline_type

        if biological and method != "none":
            raise NotImplementedError
        version = version if not biological else 6

        best_params: dict[str, dict] = {}
        excel_filename = f"Results/{folder}/Hyperparameter_optimisation_{method}_{version}.xlsx"

        with ExcelWriter(excel_filename) as writer:  # type: ignore[abstract]
            for collection in self.collections:
                if collection == "DP" and control:
                    continue
                best_params[collection] = {}
                for filename, index_names in zip(file_names, param_names):
                    if filename != "random_forest" and control:
                        continue
                    print(collection, filename)
                    file_dir = "control" if control else f"v{version}"
                    file_path = (
                        f"Results/{folder}/{collection}/{method}/" f"{file_dir}/{filename}_hyperparameters.joblib"
                    )
                    try:
                        with open(file_path, "rb") as f:
                            rf_results = load(f)[collection]
                    except FileNotFoundError:
                        continue

                    param = {}
                    for params, folds in rf_results.items():
                        aucs = []
                        for _, auc in folds.items():
                            # Biological data here would be auc["auc"].
                            # Need to make the data compatible across
                            # pipelines
                            if isinstance(auc, dict):
                                aucs.extend(auc["auc"])
                            else:
                                aucs.extend(auc)
                        param[params] = np.mean(aucs), np.std(aucs)

                    param = {k: v for k, v in sorted(param.items(), key=lambda item: item[1][0], reverse=True)}
                    data = pd.DataFrame(data=param, index=["Mean", "STD"]).T
                    data.index.names = index_names
                    data.to_excel(writer, sheet_name=f"{collection}_{filename}")
                    best_params[collection][filename] = list(param.keys())[0], data

        wb = load_workbook(f"Results/{folder}/Hyperparameter_optimisation_{method}_{version}.xlsx")
        sheets = wb.sheetnames
        red_font = Font(color="00FF0000", bold=True)

        for sheet in sheets:
            current_sheet: Worksheet = wb[sheet]  # type: ignore
            max_col = current_sheet.max_column
            max_row = current_sheet.max_row

            rows = [
                row
                for row in current_sheet.iter_rows(min_row=1, max_col=max_col, max_row=max_row, values_only=True)
                if row[0] is not None
            ]
            max_auc = max(rows[1:], key=lambda x: x[-2])
            index = rows.index(max_auc)
            for cell in current_sheet[index + 1]:
                cell.font = red_font
        wb.save(f"Results/{folder}/Hyperparameter_optimisation_{method}_{version}.xlsx")

        self.best_params = best_params
        return best_params

    # full_protein_sets come from best_protein_set_max()
    def statistical_indicies(
        self, full_protein_sets: dict[str, dict], results_path: str, biological: bool = False
    ) -> dict[str, dict[str, dict[str, dict[str, Any]]]]:
        """Statistical inference on the feature elimination data.

        A T-test is performed on the recursive feature elimination
        protein data to identify which proteins are not statistically
        significant by the AUC of each set of proteins.

        Parameters
        ----------
        full_protein_sets
            All protein sets from recursive feature elimination.
        results_path
            The RFE results path.
        biological
            To use the biological pipeline, by default False

        Returns
        -------
            Protein data which identifies proteins to use as being
            statistically significant.
        """
        rfe_not_statistically_independent: dict[str, dict] = {}

        for collection in self.collections:
            rfe_not_statistically_independent[collection] = {}
            result = full_protein_sets[collection]
            for kde in self.kde_methods:
                if biological:
                    class_, _ = self._best_classes(result)[collection][kde]
                    rfe = result[class_][kde][-2]
                    max_auc_num = len(result[class_][kde][0])
                    with open(
                        f"Results/biological/{collection}/{kde}/rfe/{class_}.joblib",
                        "rb",
                    ) as f:
                        rfe_data_ = load(f)
                        rfe_data: dict[str, dict[int, dict[str, dict[str, list[float]]]]] = {}
                        for col, rfe_dict in rfe_data_.items():
                            rfe_data[col] = {}
                            for rfe_num, folds in rfe_dict.items():
                                rfe_data[col][rfe_num] = {}
                                aucs = []
                                for fold in folds:
                                    if fold == "proteins":
                                        continue
                                    aucs.extend(folds[fold]["aucs"])

                                rfe_data[col][rfe_num][kde] = {"aucs": aucs}
                else:
                    try:
                        rfe = result[kde][-2]
                        max_auc_num = len(result[kde][0])
                        with open(results_path, "rb") as f:
                            rfe_data = load(f)
                    except (KeyError, FileNotFoundError):
                        continue

                rfe_results = {}
                for num, folds in rfe_data[collection].items():
                    aucs = []
                    aucs.extend(folds[kde]["aucs"])
                    rfe_results[num] = aucs

                top_rfe_results = {k: rfe_results[k] for k in range(max_auc_num - 1, 0, -1)}
                top_rfe_result = rfe_results[max_auc_num]
                combos = []
                for _, value in top_rfe_results.items():
                    combos.append((top_rfe_result, value))

                # T-test of null hypothesis
                def ttest(
                    combination_values: list[tuple[list[Any], list[Any]]],
                    t_type: Literal["pairwise", "unpaired"] = "pairwise",
                ) -> tuple[int, list[str]]:
                    p_values = []
                    for combo in combination_values:
                        top, next_top = combo
                        _, p_value = (
                            ttest_rel(top, next_top, alternative="greater")
                            if t_type == "pairwise"
                            else ttest_ind(top, next_top, alternative="greater", equal_var=False)
                        )
                        p_values.append(p_value)

                    # rej = True, reject the hypothesis
                    rej, p_corrected = fdrcorrection(p_values)
                    indicies = [index for index, (r, _) in enumerate(zip(rej, p_corrected)) if not r]
                    size_model = (max_auc_num - max(indicies) - 1) if indicies else max_auc_num

                    proteins = []
                    for rfe_key, rfe_value in rfe.items():
                        if rfe_key > size_model:
                            continue
                        proteins.append(rfe_value[-1])

                    return size_model, proteins

                best_min_size_model, proteins = ttest(combos)
                best_min_size_model_ind, proteins_ind = ttest(combos, t_type="unpaired")

                dict_kde_key = f"{class_}_{kde}" if biological else kde  # type: ignore
                rfe_not_statistically_independent[collection][dict_kde_key] = {
                    "pair": {
                        "max_auc_num": max_auc_num,
                        "best_min_size_model": best_min_size_model,
                        "proteins": proteins,
                    },
                    "independent": {
                        "max_auc_num": max_auc_num,
                        "best_min_size_model": best_min_size_model_ind,
                        "proteins": proteins_ind,
                    },
                }

        return rfe_not_statistically_independent

    def kde_combinations(self, extra: dict[str, dict]) -> dict[str, dict[int | str, Proteins]]:
        """Creates protein combinations by kernel density augmentation.

        Parameters
        ----------
        extra
            The data used to create the combinations.

        Returns
        -------
            A combination of collections and proteins.
        """
        combos_collection = {}
        for collection, results in extra.items():
            data_dict: dict[int | str, list[str]] = {
                4: [],
                "a4": [],
                3: [],
                "a3": [],
                2: [],
                "a2": [],
                1: [],
                "a1": [],
                "unique_a1": [],
            }
            for i in range(4, 0, -1):
                for item in list(combinations(results.keys(), i)):
                    first = results[item[0]]["independent"]["proteins"]
                    rest = [results[collection]["independent"]["proteins"] for collection in item[1:]]
                    collection_interesct = set(first).intersection(*rest)
                    if collection_interesct:
                        temp = []
                        for protein in collection_interesct:
                            temp.append(protein)
                        data_dict[i].append((*item, temp))  # type: ignore[arg-type]
                        data_dict[f"a{i}"].extend(temp)

            data_dict["unique_a1"] = list(set(data_dict["a1"]))
            combos_collection[collection] = data_dict

        return combos_collection

    def protein_mapping(self, data: dict[str, dict[str, list[str]]] | None = None) -> dict[str, dict]:
        """Checks if any proteins are in the biological clusters.

        Parameters
        ----------
        data, optional
            A set of protein features, by default None

        Returns
        -------
            A mapping of proteins to clusters.
        """
        if data is None:
            with open("Results/RFE_feature_sets_1std.joblib", "rb") as f:
                data = load(f)
        # data = load("Results/BIOLOGICAL_RFE_feature_sets_1std.joblib")
        with open("KEGG/cluster_proteins.joblib", "rb") as f:
            cluster_proteins = load(f)

        assert data is not None

        protein_map: dict[str, dict] = {}
        for (collection, non), (_, clusters) in zip(data.items(), cluster_proteins.items()):
            protein_map[collection] = {}
            for kde in self.kde_methods:
                if kde not in non:
                    continue
                non_proteins = non[kde]
                for cluster_num, cluster in clusters.items():
                    protein_map[collection][cluster_num] = {}
                    for c_name, c_prots in cluster.items():
                        protein_map[collection][cluster_num][c_name] = []
                        for non_prot in non_proteins:
                            if non_prot in c_prots:
                                protein_map[collection][cluster_num][c_name].append(non_prot)

                        if not protein_map[collection][cluster_num][c_name]:
                            del protein_map[collection][cluster_num][c_name]
        return protein_map

    # extra comes from statistical_indicies
    def max_auc_to_excel(
        self,
        full_protein_sets: dict[str, dict],
        excel_path: str,
        collection: str,
        extra: dict[str, dict] | None = None,
        biological: bool = False,
    ) -> list[list[str | int | list[str]]]:
        """Formats and adds extra information to the Excel Workbook.

        Parameters
        ----------
        full_protein_sets
            A mapping of proteins
        excel_path
            The excel path to save data to.
        collection
            The name of the collection
        extra
            More data used to highlight statistical differences,
            by default None
        biological
            Uses the biological pipeline if True, by default False

        Returns
        -------
            A 2D list of the sheet name, the number of smallest
            proteins and a stringified list of the smallest proteins.
        """
        wb = load_workbook(excel_path)
        sheets = wb.sheetnames
        red_font = Font(color="00FF0000")
        green_font = Font(color="00009900", bold=True)
        blue_font = Font(color="00000099", bold=True)
        orange_font = Font(color="00FFBF00", bold=True)

        extras = []
        for sheet in sheets:
            if biological:
                splits = sheet.split("_")
                num, kde = int(splits[0]), "_".join(splits[1:])
                current_protein_set = full_protein_sets[collection][num][kde]
            else:
                current_protein_set = full_protein_sets[collection][sheet]
            current_sheet: Worksheet = wb[sheet]  # type: ignore
            max_col = current_sheet.max_column
            max_row = current_sheet.max_row

            dims: dict[str, int] = {}
            for row in current_sheet.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
            for col, value in dims.items():
                current_sheet.column_dimensions[col].width = value + 5

            smallest_proteins = []
            for i in range(1, current_protein_set[-1] + 1):
                protein = current_protein_set[2][i][-1]
                smallest_proteins.append(protein)

            extras.append([sheet, len(smallest_proteins), smallest_proteins])

            current_sheet.cell(row=max_row + 2, column=1).value = "Max"  # type: ignore
            current_sheet.cell(row=max_row + 3, column=1).value = len(current_protein_set[0])  # type: ignore
            cell_to_merge = current_sheet.cell(row=max_row + 2, column=2)
            cell_to_merge.value = ", ".join(current_protein_set[0])  # type: ignore
            current_sheet.merge_cells(
                start_row=max_row + 2,
                start_column=2,
                end_row=max_row + 6,
                end_column=10,
            )
            cell_to_merge.alignment = Alignment(wrap_text=True, vertical="center")  # type: ignore

            rows = [
                row
                for row in current_sheet.iter_rows(min_row=1, max_col=max_col, max_row=max_row, values_only=True)
                if row[0] is not None
            ]
            removals = list(filter(lambda x: x[0] in current_protein_set[1], rows))
            # Outside of best_auc ± 1 STD
            for removal in removals:
                idx = rows.index(removal) + 1
                for cell in current_sheet[idx + 1]:
                    cell.font = red_font

            max_auc = max(rows, key=lambda x: x[1])
            index = rows.index(max_auc) + 1
            for cell in current_sheet[index + 1]:
                cell.font = green_font

            if extra:
                if biological:
                    if sheet not in extra[collection].keys():
                        continue
                    current_sheet.sheet_properties.tabColor = "00FF00"
                for ttest_type, values in extra[collection][sheet].items():
                    best_min_size_model = values["best_min_size_model"]
                    best_auc = [row for row in rows if row[0] == best_min_size_model][0]
                    index = rows.index(best_auc) + 1
                    for cell in current_sheet[index + 1]:
                        cell.font = blue_font if ttest_type == "pair" else orange_font

                    if ttest_type == "pair":
                        continue

                    best_min_size_model_proteins = values["proteins"]
                    current_sheet.cell(row=max_row + 8, column=1).value = "Best"  # type: ignore
                    current_sheet.cell(row=max_row + 9, column=1).value = len(  # type: ignore
                        best_min_size_model_proteins
                    )
                    cell_to_merge = current_sheet.cell(row=max_row + 8, column=2)
                    cell_to_merge.value = ", ".join(best_min_size_model_proteins)  # type: ignore
                    current_sheet.merge_cells(
                        start_row=max_row + 8,
                        start_column=2,
                        end_row=max_row + 12,
                        end_column=10,
                    )
                    cell_to_merge.alignment = Alignment(wrap_text=True, vertical="center")  # type: ignore

        wb.create_sheet("Smallest_protein_sets")
        ws: Worksheet = wb["Smallest_protein_sets"]  # type: ignore
        ws.append(["Method", "# Proteins", "Proteins"])
        for info in extras:
            if isinstance(info[-1], list):
                info[-1] = ", ".join(info[-1])  # type: ignore[assignment]
            ws.append(info)

        if extra:
            wb.create_sheet("Combinations")
            ws: Worksheet = wb["Combinations"]  # type: ignore
            ws.append([collection])
            combos_collection = self.kde_combinations(extra)
            for kde_value, values in combos_collection[collection].items():
                items = []
                for info in values:
                    if not isinstance(kde_value, str):
                        items.append(
                            [
                                ", ".join([*info[:-1]]),
                                len(info[-1]),
                                ", ".join(info[-1]),
                            ]
                        )
                if isinstance(kde_value, str):
                    items.append([kde_value, len(values), ", ".join(values)])
                for item in items:
                    ws.append(item)

        wb.save(excel_path)
        return extras

    def get_features(self, importances: list[np.floating], selections: list) -> tuple[list, list]:
        """Gets features from a list of proteins.

        Sorts the features from selections based on the given
        feature importance scores. Used for recursive feature elimination.

        Parameters
        ----------
        importances
            The feature importance scores for each feature
        selections
            The current list of proteins to go through feature selection

        Returns
        -------
            The next set of proteins and the proteins to remove.
        """
        top_features = np.argsort(importances)[::-1]
        to_remove = list(selections)[top_features[-1]]
        selections.remove(to_remove)

        return selections, to_remove

    def get_feature_importances(self, results: dict[int, dict]) -> list[np.floating]:
        """Finds the averaged feature importance scores.

        Parameters
        ----------
        results
            The folds which hold auc and feature importance data

        Returns
        -------
            The averaged feature importance scores for each feature
        """
        results = {k: v for k, v in results.items() if k != "proteins"}
        chunked = [chunk[i]["feature_importances"] for i, chunk in zip(range(len(results)), self._chunk_dict(results))]

        feature_importances = []
        for chunk in zip(*chunked):
            feature_importances.append(np.mean(chunk))

        return feature_importances

    @abstractmethod
    def best_protein_set_max(self, path: str, results_path: str) -> dict[str, dict]:  # noqa: U100
        """Calculates the best set of proteins to use.

        All data is exported to an Excel sheet which includes
        all of the collections, augmentation methods and the calculated
        auc values.

        Parameters
        ----------
        path
            The excel path to save data to.
        results_path
            The RFE results path.

        Returns
        -------
            A comprehensive mapping of proteins, recursive
            elimination values and other metadata, keyed by
            collection and augmentation method.
        """
        ...

    def get_final_protein_sets(self, results_path: str, all_models: bool = False) -> dict[str, list[str]]:
        """Gets the test set proteins for each collection.

        Parameters
        ----------
        results_path
            The path to the results
        all_models
            If True, then find proteins per model, by default False

        Returns
        -------
            The final set of proteins for each collection.
        """
        biological_flag = True if self.combo_options.pipeline_type == "biological" else False
        suffix = "B" if biological_flag else "NB"
        folder = "biological" if biological_flag else "non_biological"
        if all_models:
            file_path = f"Results/{folder}/{self.collection}/Final_protein_sets_{suffix}_{self.model_name}.joblib"
        else:
            file_path = f"Results/{folder}/Final_protein_sets_{suffix}.joblib"

        try:
            with open(file_path, "rb") as file:
                data: dict[str, list[str]] = load(file)
        except FileNotFoundError:
            excel_path = (
                f"Results/{folder}/{self.collection}/{self.kde_method}/"
                f"{self.collection}_{self.model_name}_best_params.xlsx"
            )
            full_protein_sets = self.best_protein_set_max(excel_path, results_path)
            stats = self.statistical_indicies(full_protein_sets, results_path, biological=biological_flag)
            kde_data = self.kde_combinations(stats)
            data = {}
            for collection in self.collections:
                data[collection] = kde_data[collection]["unique_a1"]

            with open(file_path, "wb") as path:
                dump(data, path)

        return data

    def run_grid_search(
        self,
        df: DataFrame,
        combo_options: data_models.ComboOptions,
        params: Sequence[Parameter] | None = None,
        rfe: bool = False,
        model: type[BaseEstimator] = RandomForestClassifier,
    ) -> list[tuple[dict[str, Any], dict[str, Any]]]:
        """Loops through all MCCV folds for each set of hyperparameters.

        Parameters
        ----------
        df
            The original data
        combo_options
            The set of values used for this combination
        params
            The set of hyperparameters to use. If None, all defaults will
            be used, default None
        rfe
            If True, then training is altered for recursive feature
            elimination, default False
        model
            The type of estimator to use, default RandomForestClassifier

        Returns
        -------
            Data ordered by hyperparameters with AUC scores.
        """
        if params is None:
            params = [Parameter(model)]

        model_dict: dict[BaseEstimator, ModelEstimator] = {
            RandomForestClassifier: RandomForest(df, combo_options, self.constant_options, params),
            LogisticRegression: LogReg(df, combo_options, self.constant_options, params),
            SVC: SupportVectorMachine(df, combo_options, self.constant_options, params),
            MLPClassifier: MultiLayerPerceptron(df, combo_options, self.constant_options, params),
            DecisionTreeClassifier: DecisionTree(df, combo_options, self.constant_options, params),
            XGBClassifier: XGBoost(df, combo_options, self.constant_options, params),
        }  # type: ignore

        class_model = model_dict[model]  # type: ignore
        kde_name = cast(KDEType, combo_options.kde_method.name.lower())
        class_model.train_info[kde_name] = {}

        options = {
            class_model.preprocessor: True,
            combo_options.kde_method.name.lower(): combo_options.kde_method.value,
        }

        # Use joblib to parellilze for faster processing
        print(f"------------------------{combo_options.collection}------------------------")
        best_params = Parallel(n_jobs=-1, verbose=50)(
            delayed(self.train_model_parameters)(parameters, class_model, df, options, rfe)
            for parameters in class_model.hyperparameters
        )
        class_model.flush(kde_name)

        return best_params

    def train_model_parameters(
        self,
        hyperparameters: dict[str, Any],
        model: ModelEstimator,
        data: DataFrame,
        options: dict[str, Any],
        rfe: bool,
    ) -> tuple[dict[str, Any], dict[str, Any]]:
        """Trains model with specific hyperparameters.

        This can be used directly with a loop or parallelized.

        Parameters
        ----------
        hyperparameters
            A dictionary of hyperparameters for the model.
        model
            The model instance.
        data
            The data to train on.
        options
            Additional preprocessing options.
        rfe
            If True, then training is altered for recursive feature
            elimination

        Returns
        -------
            Data ordered by hyperparameters with AUC scores.
        """
        print(hyperparameters)
        # Iterate through Monte Carlo split
        for i, (train_set, validation_set) in enumerate(
            mc_cv.split(data, y=data.index.get_level_values(level="classification"))
        ):
            frames, x_valid, y_valid = model.preprocess(train_set, validation_set, **options)
            if rfe:
                model.train_rfe(frames, x_valid, y_valid, i, **hyperparameters)
            else:
                model.train(frames, x_valid, y_valid, i, **hyperparameters)

        return hyperparameters, model.train_info

    def run_pipeline(
        self, file_path: str, optimise_params: bool = False
    ) -> None | dict[str, Any] | list[dict[str, Any]]:
        """Runs a specific pipeline.

        Parameters
        ----------
        file_path
            The path to load the proteins
        optimise_params, optional
            To optimise hyperparameters, by default False

        Raises
        ------
        ValueError
            `optimise_params` should not be True if the
            hyperparameters are default.
        """
        np.random.seed(self.constant_options.constants.RANDOM_STATE.value)
        collection = self.collection
        kde_method = self.kde_method
        pipeline_stage = self.constant_options.stage.value

        if not optimise_params:
            save_path = f"Results/{self.pipeline}/{collection}/{kde_method}"
            save_path_end = f"{self.model_name}_{pipeline_stage}"
            if (len(self.hyperparameters) > 1) or self.hyperparameters[0].name != "default":
                raise ValueError("Default values should be used if not optimizing for model hyperparameters.")
        else:
            save_path = f"Results/{self.pipeline}/{collection}/{kde_method}/v{self.constant_options.version}"
            save_path_end = f"{self.model_name}_hyperparameters"

        results: dict[str, Any] = {collection: {}}
        print(f"Running {pipeline_stage=}")
        if pipeline_stage == "validation":
            gridsearch = self._run_validation(file_path)
            for (parameters, kde_data) in gridsearch:
                for _, folds in kde_data.items():
                    if self.constant_options.feature_selection:
                        results[collection] = {kde_method: folds}
                    else:
                        results[collection][tuple(parameters.values())] = folds
            if self.constant_options.feature_selection:
                results = [results]  # type: ignore
        elif pipeline_stage == "testing":
            results = self._run_testing(file_path)
        else:  # RFE
            results = self._run_recursive_feature_elimination(file_path)

        os.makedirs(save_path, exist_ok=True)
        with open(f"{save_path}/{save_path_end}.joblib", "wb") as dump_data:
            print(f"Results saved at {save_path}/{save_path_end}.joblib")
            dump(results, dump_data)

        return results

    def _run_validation(self, file_path: str) -> list[tuple[dict[str, Any], dict[str, Any]]]:
        data = self.combo_options.data.training

        if self.constant_options.feature_selection:  # type: ignore[attr-defined]
            file_contents: Proteins = data.columns.tolist()
        else:
            with open(file_path, "rb") as file:
                file_contents: Proteins = load(file)  # type: ignore

        df = data[file_contents]

        model: type[BaseEstimator] = next(iter(self.hyperparameters)).model
        gridsearch = self.run_grid_search(df, self.combo_options, params=self.hyperparameters, model=model)

        return gridsearch

    def _run_testing(self, file_path: str) -> dict[str, float]:
        collection = self.combo_options.collection
        kde_method = cast(KDEType, self.combo_options.kde_method.name.lower())
        partition = self.combo_options.data

        with open(file_path, "rb") as f:
            selections: Proteins = load(f)[collection]

        train_df, test_df = partition.training, partition.testing
        train_df, test_df = train_df[selections], test_df[selections]

        best_params = self.get_best_hyperparameters(
            kde_method,
            control=self.constant_options.control_data,
            biological=self.constant_options.biological,
            version=self.constant_options.version,
        )

        current_best_params = list(best_params[collection][self.model_name][0])
        current_best_params = [
            p if not isinstance(p, float) else None if np.isnan(p) else p for p in current_best_params
        ]
        current_best_param_names = best_params[collection][self.model_name][1].index.names

        params = [
            Parameter(RandomForestClassifier, name, [value])
            for name, value in zip(current_best_param_names, current_best_params)
        ]
        self.hyperparameters = params

        rf_model = RandomForest(train_df, self.combo_options, self.constant_options, params=params)
        rf_model.clf.set_params(**next(iter(rf_model.hyperparameters)))

        targets_train = train_df.index.get_level_values(level="classification").astype("category")
        targets_test = test_df.index.get_level_values(level="classification").astype("category")

        train = pd.concat(
            [
                pd.Series(targets_train),
                pd.DataFrame(train_df.values, columns=train_df.columns),
            ],
            axis=1,
        )
        testing = pd.concat(
            [
                pd.Series(targets_test),
                pd.DataFrame(test_df.values, columns=test_df.columns),
            ],
            axis=1,
        )
        x_test = testing.drop("classification", axis=1)
        y_test = testing["classification"].astype("category")

        frames = rf_model.augment(
            train,
            {
                self.combo_options.preprocessor: True,
                kde_method: self.combo_options.kde_method.value,
            },
        )
        auc = rf_model.train_testing(frames, x_test, y_test)

        return {collection: auc}

    def _run_recursive_feature_elimination(self, file_path: str) -> dict[str, dict[int, dict]]:
        collection = self.collection
        kde_method = self.kde_method

        with open(file_path, "rb") as file:
            selections = load(file)[collection]

        param_values, param_data = self.get_best_hyperparameters(
            kde_method,
            control=self.constant_options.control_data,
            biological=self.constant_options.biological,
            version=self.constant_options.version,
        )[collection][self.model_name]

        param_names = param_data.index.names

        params = [Parameter(RandomForestClassifier, name, [value]) for name, value in zip(param_names, param_values)]

        rfe_results = self._run_rfe_loop(collection, kde_method, params, proteins=selections)

        return rfe_results

    def _run_rfe_loop(
        self,
        collection: str,
        kde_method: KDEType,
        params: Sequence[Parameter],
        *,
        proteins: list[str],
    ) -> dict[str, dict[int, dict]]:
        rfe_results: dict[str, dict[int, dict]] = {collection: {}}
        while len(proteins) > 0:
            df = self.combo_options.data.training[list(proteins)]

            model = next(iter(self.hyperparameters)).model
            gridsearch_score = self.run_grid_search(df, self.combo_options, params=params, rfe=True, model=model)

            results = {}
            for (_, data) in gridsearch_score:
                for kde, folds in data.items():
                    results[kde] = folds

            iteration = len(proteins)
            print(f"# Features: {iteration}, Features: {proteins}")
            proteins, removed = self.get_features(self.get_feature_importances(results[kde_method]), proteins)

            aucs = []
            for _, props in results[kde_method].items():
                if "feature_importances" in props:
                    del props["feature_importances"]

                if "auc" in props:
                    aucs.extend(props["auc"])

            results = {kde_method: {"protein_removed": removed, "aucs": aucs}}
            rfe_results[collection][iteration] = results

        return rfe_results

    # Biological only
    def _best_classes(self, result: dict[int, dict]) -> dict[str, dict]:
        """Generates the best classes of the 16 biological configurations.

        Parameters
        ----------
        result
            The results of the best protein set

        Returns
        -------
            The result dict with only the best biological classification.
        """
        best_classes: dict[str, dict] = {}
        for collection in self.collections:
            best_classes[collection] = {}
            for kde in self.kde_methods:
                highest = 0
                for i in range(1, 17):
                    rfe = result[i][kde][2]
                    smallest_within_std = result[i][kde][3]
                    best_rfe: list = rfe[smallest_within_std]
                    if highest < best_rfe[0]:
                        highest = best_rfe[0]
                        best_classes[collection][kde] = (i, (smallest_within_std, *best_rfe))
        return best_classes

    def _chunk_dict(self, data: dict[int, dict], size: int = 1) -> Generator:
        """Yields chunks of data containing feature importance scores.

        Parameters
        ----------
        data
            The folds which hold auc and feature importance data
        size
            The stopping index for islice, by default 1

        Yields
        ------
            Data used in chunking the feature importance scores.
        """
        it = iter(data)
        for _ in range(0, len(data), size):
            yield {k: data[k] for k in islice(it, size)}

    def _create_best_protein_dataframe(
        self,
        rfe: dict[int, tuple[float, float, float, str]],
        writer: ExcelWriter,
        sheet_name: str,
    ) -> tuple[set[int], int]:
        """Creates additional columns and exports to an excel file.

        Parameters
        ----------
        rfe
            The dictionary containing the recursive feature
            elimination cleaned data
        writer
            The ExcelWriter object
        sheet_name
            The sheet name

        Returns
        -------
            Indicies to be removed and the best index of `rfe`,
            computed by max AUC.
        """
        best_auc, best_std = (
            rfe[max(rfe, key=rfe.get)][0],  # type: ignore[arg-type]
            rfe[max(rfe, key=rfe.get)][1],  # type: ignore[arg-type]
        )
        best_variable_set = min(k for k, v in rfe.items() if v[0] == best_auc)

        lower = best_auc - best_std
        upper = best_auc + best_std
        rfe_removal = {k for k, v in rfe.items() if lower > v[0] < upper}

        rfe_df = pd.DataFrame(rfe, index=Index(("Mean", "STD", "SEM", "Proteins"))).T
        rfe_df["Mean (2DP)"] = rfe_df["Mean"].apply(lambda x: round(x, 2))
        rfe_df["Mean - STD"] = rfe_df["Mean"] - rfe_df["STD"]
        rfe_df["Mean - SEM"] = rfe_df["Mean"] - rfe_df["SEM"]
        rfe_df["STD (2DP)"] = rfe_df["STD"].apply(lambda x: round(x, 2))
        rfe_df["SEM (2DP)"] = rfe_df["SEM"].apply(lambda x: round(x, 2))
        rfe_df["Mean (2DP) - STD (2DP)"] = rfe_df["Mean (2DP)"] - rfe_df["STD (2DP)"]
        rfe_df["Mean (2DP) - SEM (2DP)"] = rfe_df["Mean (2DP)"] - rfe_df["SEM (2DP)"]
        rfe_df = rfe_df.reindex(
            [
                "Mean",
                "Mean (2DP)",
                "STD",
                "STD (2DP)",
                "SEM",
                "SEM (2DP)",
                "Mean - STD",
                "Mean - SEM",
                "Mean (2DP) - STD (2DP)",
                "Mean (2DP) - SEM (2DP)",
                "Proteins",
            ],
            axis=1,
        )
        rfe_df.to_excel(writer, sheet_name=sheet_name)

        return rfe_removal, best_variable_set

    def _find_proteins_within_std(
        self, rfe: dict[int, tuple[float, float, float, str]], best_variable_set: int
    ) -> Proteins:
        """Accumulates all proteins from the given `best_variable_set`.

        Parameters
        ----------
        rfe
            The dictionary containing the recursive feature
            elimination cleaned data
        best_variable_set
            The best index of `rfe`, computed by max AUC

        Returns
        -------
            A list of the best proteins.
        """
        proteins = []
        for rfe_key, rfe_data in rfe.items():
            if rfe_key > best_variable_set:
                continue
            proteins.append(rfe_data[-1])

        return proteins
